import glob
import math
import os
from openpyxl.styles import Font
import cv2
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
import utils
import matplotlib.patches as patches
from osgeo import gdal_array as ga


def cal_num(image_path, model, device, size=15, output_dir=None):
    model = model.to(device)
    image_list = glob.glob(os.path.join(image_path + '*.tif'))
    for i in range(len(image_list)):

        base_name = os.path.splitext(os.path.basename(image_list[i]))[0] + '.png'
        print(base_name)
        raster0, mask_arr, raster = utils.crop_image_and_pre(image_list[i], model, device)
        kernel = np.ones((size, size), np.uint8)
        opened_image = cv2.morphologyEx(raster, cv2.MORPH_OPEN, kernel)
        labeled = utils.find_connected_components(opened_image, direction_num=8)
        label_num_list = []
        fig, ax = plt.subplots()
        ax.imshow(opened_image, cmap='gray')
        ax.axis('off')
        properties = utils.calculate_region_properties(labeled)
        for label, props in properties.items():
            if props['area'] >= 500:
                label_num_list.append(label)
                rectangle = patches.Rectangle((props['bbox'][0], props['bbox'][2]), props['bbox_width'],
                                              props['bbox_height'],
                                              linewidth=1, edgecolor='r', facecolor='none')
                ax.add_patch(rectangle)
        print('Num', len(label_num_list))
        plt.show()
        if output_dir is not None:
            original_output_dir = os.path.join(output_dir, "original")
            opened_output_dir = os.path.join(output_dir, "opened")
            os.makedirs(original_output_dir, exist_ok=True)
            os.makedirs(opened_output_dir, exist_ok=True)
            ga.SaveArray(raster, os.path.join(original_output_dir, base_name))
            ga.SaveArray(opened_image, os.path.join(opened_output_dir, base_name))

    return raster, opened_image


def cal_fauna_length_width(image_path, model, device, window_size=11, max_window_size=25, look_back=30,
                           window_size2=20,
                           excel_output_path='numbers.xlsx', save_folder=r'E:\\results\\',
                           vis_skeleton=False, visualization=False):
    scale = utils.cal_scale(image_path)
    columns = ['Name', 'Count', 'Label', 'Area', 'position_x1', 'position_x2', 'position_y1',
               'position_y2', 'scale(pixel/mm)', 'length(pixel)', 'width(pixel)', 'pixel_length', 'pixel_width',
               'length(mm)', 'width(mm)', 'status']
    animal_data = pd.DataFrame(columns=columns)
    base_name = os.path.splitext(os.path.basename(image_path))[0] + '.png'

    image_raster, raster_origin, raster_opened, properties_origin, properties_opened = utils.find_rect_new(image_path,
                                                                                                           model,
                                                                                                           device,
                                                                                                           size=window_size,
                                                                                                           max_size=max_window_size)
    area_list = [v['area'] for v in properties_opened.values()]
    area_list = np.array(area_list)
    num = 0
    # fig, ax = plt.subplots()
    # print(area_list)
    # ax.imshow(raster_opened, cmap='gray')
    # ax.axis('off')
    position_list = []
    label_num = 0
    print('area_list', area_list)
    data_list = []
    length_list = []
    error_list = []
    fig, ax = plt.subplots()
    ax.imshow(raster_opened, cmap='gray')
    ax.axis('off')
    save_name = save_folder + os.path.splitext(os.path.basename(image_path))[0] + '-binary.png'
    for label, props in properties_opened.items():
        if props['area'] >= np.mean(area_list) / 4:
            label_num += 1
    for label, props in properties_opened.items():
        if props['area'] >= np.mean(area_list) / 4:
            num += 1
            image_raster_path = save_folder + os.path.splitext(os.path.basename(image_path))[
                0] + '-image-' + str(num) + '.tif'
            image_skeleton_name = save_folder + os.path.splitext(os.path.basename(image_path))[
                0] + '-skeleton-' + str(num) + '.tif '
            image_extended_skeleton_name = save_folder + \
                                           os.path.splitext(os.path.basename(image_path))[
                                               0] + '-extended-skeleton-' + str(num) + '.tif'
            image_length_width_name = save_folder + os.path.splitext(os.path.basename(image_path))[
                0] + '-length-width-' + str(num) + '.tif'
            rectangle = patches.Rectangle((props['bbox'][0], props['bbox'][2]), props['bbox_width'],
                                          props['bbox_height'],
                                          linewidth=1, edgecolor='r', facecolor='none')
            ax.add_patch(rectangle)
            position_list.append([props['bbox'][2], props['bbox'][3], props['bbox'][0], props['bbox'][1]])
            # print('position_list', position_list)
            image_arr = image_raster[props['bbox'][2]:props['bbox'][3] + 1, props['bbox'][0]:props['bbox'][1] + 1, :]
            image_arr1 = image_arr.transpose(2, 0, 1)
            print('image_arr', image_arr.shape)
            ga.SaveArray(image_arr1, image_raster_path)

            arr = raster_opened[props['bbox'][2]:props['bbox'][3] + 1, props['bbox'][0]:props['bbox'][1] + 1]
            filter_arr = utils.find_rect_and_remove(arr)
            plt.subplot(131)
            plt.title('origin iamge')
            plt.imshow(image_arr)
            plt.axis('off')
            plt.subplot(132)
            plt.title('arr')
            plt.imshow(arr, cmap='gray')
            plt.axis('off')
            # plt.show()
            plt.subplot(133)
            plt.title('filter_arr')
            plt.imshow(filter_arr, cmap='gray')
            plt.axis('off')
            plt.show()

            longest_skeleton, path, smoothed_longest_skeleton, smoothed_path, extended_skeleton, extended_path, geo_length, pixel_length = cal_fauna_major_axis(
                filter_arr, look_back=look_back)
            ga.SaveArray(smoothed_longest_skeleton, image_skeleton_name)
            ga.SaveArray(extended_skeleton, image_extended_skeleton_name)
            if vis_skeleton:
                # plt.title('filter_arr')
                # plt.imshow(filter_arr, cmap='gray')
                # plt.show()
                plt.subplot(121)
                plt.title('smoothed skeleton')
                plt.axis('off')
                plt.imshow(smoothed_longest_skeleton, cmap='gray')
                # plt.show()
                plt.subplot(122)
                plt.title('extended skeleton')
                plt.axis('off')
                plt.imshow(extended_skeleton, cmap='gray')
                plt.show()
            print('Area', props['area'])
            print(f"Length：{geo_length} pixel")

            best_width, best_pos_point, best_neg_point = utils.cal_fauna_minor_axis(filter_arr, extended_path, 0,
                                                                              size=window_size2)
            # print('best_pos_point', best_pos_point, 'best_neg_point', best_neg_point)
            width = math.sqrt(
                np.power((best_pos_point[0] - best_neg_point[0]), 2) + np.power((best_pos_point[1] - best_neg_point[1]),
                                                                                2))
            length_width_arr = utils.visualize_axis(filter_arr, extended_path, best_pos_point, best_neg_point)
            ga.SaveArray(length_width_arr, image_length_width_name)
            print(f"Width：{width} pixel")
            geo_length = round(geo_length, 2)
            width = round(width, 2)
            geo_length_scale = geo_length / 1000
            geo_width_scale = width / 1000
            geo_length_scale = round(geo_length_scale, 4)
            geo_width_scale = round(geo_width_scale, 4)
            length_list.append(geo_length)
            data_list.append(
                [num, props['area'], props['bbox'][0], props['bbox'][1], props['bbox'][2], props['bbox'][3], scale,
                 geo_length, width, pixel_length, best_width, geo_length_scale, geo_width_scale])

            if visualization:
                filter_arr = utils.visualize_axis(filter_arr, extended_path, best_pos_point, best_neg_point)
                plt.imshow(filter_arr, cmap='gray')
                plt.axis('off')
                plt.show()
    # plt.savefig(save_name, dpi=300)
    median = np.median(np.array(length_list))
    for j in range(len(length_list)):
        error_list.append((length_list[j] - median) / median)
    print('length_list', length_list)
    print('error_list', error_list)
    for i in range(len(data_list)):
        if error_list[i] < 0.5:
            animal_data = pd.concat(
                [animal_data, pd.DataFrame(
                    {'Name': [base_name], 'Count': [label_num], 'Label': data_list[i][0], 'Area': data_list[i][1],
                     'position_x1': data_list[i][2], 'position_x2': data_list[i][3],
                     'position_y1': data_list[i][4], 'position_y2': data_list[i][5],
                     'scale(pixel/mm)': data_list[i][6],
                     'length(pixel)': data_list[i][7], 'width(pixel)': data_list[i][8],
                     'pixel_length': data_list[i][9], 'pixel_width': data_list[i][10],
                     'length(mm)': data_list[i][11],
                     'width(mm)': data_list[i][12], 'status': True})], ignore_index=True)
        if error_list[i] >= 0.5:
            animal_data = pd.concat(
                [animal_data, pd.DataFrame(
                    {'Name': [base_name], 'Count': [label_num], 'Label': data_list[i][0], 'Area': data_list[i][1],
                     'position_x1': data_list[i][2], 'position_x2': data_list[i][3],
                     'position_y1': data_list[i][4], 'position_y2': data_list[i][5],
                     'scale(pixel/mm)': data_list[i][6],
                     'length(pixel)': data_list[i][7], 'width(pixel)': data_list[i][8],
                     'pixel_length': data_list[i][9], 'pixel_width': data_list[i][10],
                     'length(mm)': data_list[i][11],
                     'width(mm)': data_list[i][12], 'status': False})], ignore_index=True)

    if excel_output_path is not None:
        file_exists = os.path.exists(excel_output_path)

        with pd.ExcelWriter(
                excel_output_path,
                engine="openpyxl",
                mode="a" if file_exists else "w",
                if_sheet_exists="overlay" if file_exists else None
        ) as writer:
            if file_exists:
                book = load_workbook(excel_output_path)
                if "Sheet1" not in book.sheetnames:
                    book.create_sheet("Sheet1")
                    book.save(excel_output_path)
                startrow = book["Sheet1"].max_row
            else:
                startrow = 0

            animal_data.to_excel(
                writer,
                sheet_name="Sheet1",
                startrow=startrow,
                index=False,
                header=False if startrow > 0 else True
            )

            workbook = writer.book
            worksheet = workbook["Sheet1"]
            red_font = Font(color="FF0000")
            black_font = Font(color="000000")
            red_rows = animal_data.index[animal_data['status'] == False].tolist()
            for df_row_idx in range(len(animal_data)):
                excel_row_num = startrow + df_row_idx + 1
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=excel_row_num, column=col)
                    cell.font = red_font if df_row_idx in red_rows else black_font

            workbook.save(excel_output_path)
